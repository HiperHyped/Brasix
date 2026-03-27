from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT_DIR / "dados"
DATA_DIR = ROOT_DIR / "data"
WORKBOOK_GLOB = "*Mesorregi*.xlsx"
ICONS_FILE = RAW_DIR / "icones_commodities.txt"

CATEGORY_SPECS = [
    ("agro", "Agro", "#5b8f4d", range(5, 15)),
    ("pecuaria", "Pecuaria", "#8a5b34", range(15, 20)),
    ("florestal", "Florestal", "#2d7a63", range(20, 23)),
    ("mineral", "Mineral", "#b46a2b", range(23, 32)),
    ("energia", "Energia", "#c8501e", range(32, 35)),
]

NAME_ALIASES = {
    "min ferro": "ferro",
    "pesca aquic": "pesca",
    "carvao min": "carvao",
}


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return NAME_ALIASES.get(text, text)


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def workbook_path() -> Path:
    matches = sorted(RAW_DIR.glob(WORKBOOK_GLOB))
    if not matches:
        raise FileNotFoundError(f"Nenhuma planilha encontrada em {RAW_DIR} com padrao {WORKBOOK_GLOB}.")
    return matches[0]


def load_icons() -> dict[str, dict[str, str]]:
    payload: dict[str, dict[str, str]] = {}
    for raw_line in ICONS_FILE.read_text(encoding="utf-8").splitlines():
        if ":" not in raw_line:
            continue
        left, right = raw_line.split(":", 1)
        name = left.strip()
        icon = right.strip()
        payload[normalize_key(name)] = {"name": name, "icon": icon}
    return payload


def category_for_index(index: int) -> tuple[str, str, str]:
    for category_id, category_label, color, column_range in CATEGORY_SPECS:
        if index in column_range:
            return category_id, category_label, color
    raise ValueError(f"Coluna de commodity fora do intervalo esperado: {index}")


def extract_unit(header_value: str) -> str:
    match = re.search(r"\((.*?)\)", header_value)
    return match.group(1) if match else ""


def clean_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "–", "—"}:
        return None
    text = text.replace(".", "").replace(",", ".")
    return float(text)


def build_payload() -> tuple[list[dict[str, object]], list[dict[str, object]], dict[str, float]]:
    icon_map = load_icons()
    workbook = load_workbook(workbook_path(), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))
    commodity_headers = header_row[5:35]

    commodities: list[dict[str, object]] = []
    commodity_ids: list[str] = []
    for index, header in enumerate(commodity_headers, start=5):
        header_text = str(header or "")
        stem = header_text.split("\n", 1)[0].strip()
        key = normalize_key(stem)
        icon_entry = icon_map.get(key, {"name": stem, "icon": "📦"})
        category_id, category_label, color = category_for_index(index)
        commodity_id = slugify(icon_entry["name"])
        commodity_ids.append(commodity_id)
        commodities.append(
            {
                "id": commodity_id,
                "name": icon_entry["name"],
                "category": category_id,
                "category_label": category_label,
                "unit": extract_unit(header_text),
                "icon": icon_entry["icon"],
                "color": color,
                "source_column": header_text,
            }
        )

    cities: list[dict[str, object]] = []
    seen_ids: dict[str, int] = {}
    latitudes: list[float] = []
    longitudes: list[float] = []

    for row in worksheet.iter_rows(min_row=5, values_only=True):
        if not row or row[0] is None:
            continue

        state_code = str(row[0]).strip()
        source_region_name = str(row[1]).strip()
        state_name = str(row[2]).strip()
        city_name = str(row[3]).strip()
        population = clean_number(row[4]) or 0.0
        latitude = clean_number(row[35])
        longitude = clean_number(row[36])
        if latitude is None or longitude is None:
            continue

        city_id_base = f"{state_code.lower()}-{slugify(city_name)}"
        seen_ids[city_id_base] = seen_ids.get(city_id_base, 0) + 1
        city_id = city_id_base if seen_ids[city_id_base] == 1 else f"{city_id_base}-{seen_ids[city_id_base]:02d}"

        commodity_values: dict[str, float] = {}
        for commodity_id, raw_value in zip(commodity_ids, row[5:35]):
            numeric_value = clean_number(raw_value)
            if numeric_value is None or numeric_value <= 0:
                continue
            commodity_values[commodity_id] = numeric_value

        dominant_commodity_id = None
        if commodity_values:
            dominant_commodity_id = max(commodity_values.items(), key=lambda item: item[1])[0]

        cities.append(
            {
                "id": city_id,
                "name": city_name,
                "label": f"{city_name}, {state_code}",
                "state_code": state_code,
                "state_name": state_name,
                "source_region_name": source_region_name,
                "population_thousands": population,
                "latitude": latitude,
                "longitude": longitude,
                "commodity_values": commodity_values,
                "dominant_commodity_id": dominant_commodity_id,
            }
        )
        latitudes.append(latitude)
        longitudes.append(longitude)

    map_config = {
        "center_lat": round(sum(latitudes) / len(latitudes), 4),
        "center_lon": round(sum(longitudes) / len(longitudes), 4),
        "lat_min": round(min(latitudes) - 3.0, 4),
        "lat_max": round(max(latitudes) + 3.0, 4),
        "lon_min": round(min(longitudes) - 3.0, 4),
        "lon_max": round(max(longitudes) + 3.0, 4),
    }

    return commodities, cities, map_config


def save_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def ensure_routes_file() -> None:
    routes_path = DATA_DIR / "routes.json"
    if routes_path.exists():
        return
    save_json(routes_path, {"edges": []})


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    commodities, cities, map_config = build_payload()
    save_json(DATA_DIR / "commodities.json", commodities)
    save_json(DATA_DIR / "cities.json", cities)
    save_json(DATA_DIR / "map_config.json", map_config)
    ensure_routes_file()
    print(f"commodities={len(commodities)} cities={len(cities)}")


if __name__ == "__main__":
    main()
