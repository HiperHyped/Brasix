from __future__ import annotations

import argparse
import json
import shutil
import sys
import unicodedata
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK_PATH = ROOT_DIR / "dados" / "brasix_oferta_demanda_reescritas.xlsx"
SUPPLY_OUTPUT_PATH = ROOT_DIR / "v1" / "json" / "game" / "city_product_supply_matrix.json"
DEMAND_OUTPUT_PATH = ROOT_DIR / "v1" / "json" / "game" / "city_product_demand_matrix.json"
CITY_CATALOG_PATH = ROOT_DIR / "v1" / "json" / "city_catalog.json"
PRODUCT_MASTER_PATH = ROOT_DIR / "v1" / "json" / "game" / "product_master_v1_1.json"
BACKUP_ROOT = ROOT_DIR / "v1" / "json" / "game" / "_canonical_mod_backups"

SHEET_TO_OUTPUT = {
    "oferta_reescrita": ("supply", SUPPLY_OUTPUT_PATH),
    "demanda_reescrita": ("demand", DEMAND_OUTPUT_PATH),
}

HEADER_ROW_INDEX = 4
DATA_START_ROW_INDEX = 5
PRODUCT_START_COLUMN_INDEX = 6
LATITUDE_COLUMN_INDEX = 50
LONGITUDE_COLUMN_INDEX = 51
SOURCE_COLUMN_INDEX = 52

EMPTY_MARKERS = {"", "-", "–", "—", "n/a", "na", "null", "none"}

PRODUCT_NAME_ALIASES = {
    "cana de acucar": "cana-de-acucar",
    "gas natural": "gas-natural",
    "sal marinho": "sal-marinho",
    "terras raras": "terras-raras",
    "pao": "pao",
    "lacteos": "lacteos",
    "veiculos": "veiculos",
}


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = []
    previous_was_separator = False
    for char in ascii_text:
        if char.isalnum():
            cleaned.append(char)
            previous_was_separator = False
            continue
        if not previous_was_separator:
            cleaned.append(" ")
            previous_was_separator = True
    return " ".join("".join(cleaned).split())


def product_name_from_header(header_value: Any) -> str:
    text = str(header_value or "").replace("\n", " ").strip()
    if "(" in text:
        text = text.split("(", 1)[0].strip()
    return normalize_text(text)


def build_product_lookup(master_products: list[dict[str, Any]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for product in master_products:
        product_id = str(product.get("id") or "").strip()
        if not product_id:
            continue
        names = {
            normalize_text(product.get("name")),
            normalize_text(product_id.replace("-", " ")),
            normalize_text((product.get("legacy_source_product_id") or "").replace("-", " ")),
        }
        for name in names:
            if name:
                lookup[name] = product_id
        display_name = normalize_text(product.get("short_name"))
        if display_name:
            lookup[display_name] = product_id
    for alias_name, product_id in PRODUCT_NAME_ALIASES.items():
        lookup[alias_name] = product_id
    return lookup


def build_city_lookup(city_catalog: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], str], dict[tuple[str, str], str]]:
    by_region: dict[tuple[str, str], str] = {}
    by_city: dict[tuple[str, str], str] = {}
    for city in city_catalog:
        city_id = str(city.get("id") or "").strip()
        state_code = normalize_text(city.get("state_code"))
        if not city_id or not state_code:
            continue
        region_name = normalize_text(city.get("source_region_name"))
        city_name = normalize_text(city.get("name"))
        if region_name:
            by_region[(state_code, region_name)] = city_id
        if city_name:
            by_city[(state_code, city_name)] = city_id
    return by_region, by_city


def parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    normalized = normalize_text(text)
    if normalized in EMPTY_MARKERS:
        return None
    compact = text.replace(".", "").replace(",", ".")
    compact = compact.replace(" ", "")
    try:
        return float(compact)
    except ValueError:
        return None


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def backup_existing_file(path: Path, backup_dir: Path) -> Path | None:
    if not path.exists():
        return None
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / path.name
    shutil.copy2(path, backup_path)
    return backup_path


def iter_sheet_rows(sheet) -> Iterable[tuple[int, tuple[Any, ...]]]:
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        yield row_index, row


def import_sheet(
    sheet,
    layer: str,
    output_path: Path,
    workbook_path: Path,
    product_lookup: dict[str, str],
    city_by_region: dict[tuple[str, str], str],
    city_by_name: dict[tuple[str, str], str],
) -> dict[str, Any]:
    rows = list(iter_sheet_rows(sheet))
    header_row = next((row for index, row in rows if index == HEADER_ROW_INDEX), None)
    if header_row is None:
        raise ValueError(f"Aba '{sheet.title}' sem linha de cabecalho {HEADER_ROW_INDEX}.")

    product_columns: list[tuple[int, str, str]] = []
    unknown_headers: list[str] = []
    for column_index in range(PRODUCT_START_COLUMN_INDEX - 1, LATITUDE_COLUMN_INDEX - 1):
        header_value = header_row[column_index] if column_index < len(header_row) else None
        product_name = product_name_from_header(header_value)
        if not product_name:
            continue
        product_id = product_lookup.get(product_name)
        if not product_id:
            unknown_headers.append(str(header_value))
            continue
        product_columns.append((column_index, product_id, str(header_value or "").strip()))

    if unknown_headers:
        raise ValueError(
            f"Aba '{sheet.title}' contem colunas de produto sem mapeamento: {', '.join(sorted(set(unknown_headers)))}"
        )

    imported_items: list[dict[str, Any]] = []
    imported_keys: set[tuple[str, str]] = set()
    unmatched_rows: list[str] = []

    for row_index, row in rows:
        if row_index < DATA_START_ROW_INDEX:
            continue
        state_code = str(row[0] or "").strip()
        region_name = str(row[1] or "").strip()
        largest_city = str(row[3] or "").strip()
        if not state_code and not region_name and not largest_city:
            continue
        summary_marker = normalize_text(" ".join(part for part in [state_code, region_name, largest_city] if part))
        if summary_marker in {"total brasil", "brasil total"}:
            continue

        state_key = normalize_text(state_code)
        region_key = normalize_text(region_name)
        city_key = normalize_text(largest_city)
        city_id = city_by_region.get((state_key, region_key)) or city_by_name.get((state_key, city_key))
        if not city_id:
            unmatched_rows.append(f"{state_code} | {region_name} | {largest_city}")
            continue

        for column_index, product_id, _header in product_columns:
            value = parse_numeric(row[column_index] if column_index < len(row) else None)
            if value is None or value == 0:
                continue
            item = {
                "id": f"city_product_{city_id}_{product_id}",
                "city_id": city_id,
                "product_id": product_id,
                "value": float(value),
            }
            imported_items.append(item)
            imported_keys.add((city_id, product_id))

    if unmatched_rows:
        preview = ", ".join(unmatched_rows[:10])
        raise ValueError(
            f"Aba '{sheet.title}' contem linhas sem correspondencia com city_catalog.json: {preview}"
        )

    existing_items: list[dict[str, Any]] = []
    existing_product_ids: set[str] = set()
    if output_path.exists():
        payload = load_json(output_path)
        if isinstance(payload, dict):
            existing_items = list(payload.get("items", []))
    imported_product_ids = {product_id for _, product_id, _ in product_columns}
    preserved_items = []
    for item in existing_items:
        city_id = str(item.get("city_id") or "").strip()
        product_id = str(item.get("product_id") or "").strip()
        key = (city_id, product_id)
        if key in imported_keys:
            continue
        if product_id in imported_product_ids:
            continue
        preserved_items.append(item)
        existing_product_ids.add(product_id)

    items = imported_items + preserved_items
    items.sort(key=lambda item: (item["city_id"], item["product_id"]))
    generated_at = datetime.now(timezone.utc).isoformat()
    return {
        "id": f"city_product_{layer}_matrix_v3",
        "seed_source": {
            "kind": "workbook_rewrite",
            "workbook": f"../../dados/{workbook_path.name}",
            "sheet": sheet.title,
            "generated_at": generated_at,
            "preserved_existing_products": sorted(existing_product_ids),
            "note": "Base canônica reescrita a partir da planilha Excel; edições manuais por mapa ficam em product_field_edits/product_field_baked.",
        },
        "items": items,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Reconstroi o MOD canonico de oferta/demanda a partir da planilha Excel.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help=f"Caminho da planilha Excel. Padrao: {DEFAULT_WORKBOOK_PATH}",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    city_catalog = load_json(CITY_CATALOG_PATH)
    if not isinstance(city_catalog, list):
        raise ValueError("city_catalog.json invalido.")
    master_payload = load_json(PRODUCT_MASTER_PATH)
    master_products = list(master_payload.get("products", [])) if isinstance(master_payload, dict) else []

    product_lookup = build_product_lookup(master_products)
    city_by_region, city_by_name = build_city_lookup(city_catalog)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    backup_dir = BACKUP_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S")
    results: list[tuple[str, Path, int, Path | None]] = []

    for sheet_name, (layer, output_path) in SHEET_TO_OUTPUT.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"A planilha nao contem a aba obrigatoria '{sheet_name}'.")
        payload = import_sheet(
            wb[sheet_name],
            layer=layer,
            output_path=output_path,
            workbook_path=workbook_path,
            product_lookup=product_lookup,
            city_by_region=city_by_region,
            city_by_name=city_by_name,
        )
        backup_path = backup_existing_file(output_path, backup_dir)
        save_json(output_path, payload)
        results.append((layer, output_path, len(payload["items"]), backup_path))

    print(f"Workbook importado: {workbook_path}")
    for layer, output_path, item_count, backup_path in results:
        print(f"[{layer}] {item_count} itens -> {output_path}")
        if backup_path:
            print(f"  backup: {backup_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - script entrypoint
        print(f"ERRO: {exc}", file=sys.stderr)
        raise
