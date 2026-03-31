from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT_DIR / "dados" / "brasix_demanda_v2_claude.xlsx"
CITY_CATALOG_PATH = ROOT_DIR / "v1" / "json" / "city_catalog.json"
PRODUCT_CATALOG_PATH = ROOT_DIR / "v1" / "json" / "product_catalog.json"
OUTPUT_PATH = ROOT_DIR / "v1" / "json" / "city_product_demand_matrix.json"

HEADER_ROW = 4
FIRST_DATA_ROW = 5
FIRST_PRODUCT_COLUMN = 6
LAST_PRODUCT_COLUMN = 35

SPECIAL_CITY_IDS = {
    ("MA", "Açailândia/Imperatriz"): "ma-imperatriz-regiao-acailandia",
    ("PE", "Vitória de Santo Antão"): "pe-caruaru-parte-vitoria-de-s-antao",
    ("TO", "Palmas / Araguaína"): "to-palmas-parte-araguaina",
}

HEADER_PRODUCT_IDS = {
    "soja": "soja",
    "milho": "milho",
    "cana-de-acucar": "cana-de-acucar",
    "algodao": "algodao",
    "cafe": "cafe",
    "arroz": "arroz",
    "trigo": "trigo",
    "laranja": "laranja",
    "feijao": "feijao",
    "mandioca": "mandioca",
    "bovinos": "bovinos",
    "suinos": "suinos",
    "aves": "aves",
    "leite": "leite",
    "ovinos": "ovinos",
    "papel-celulose": "papel-celulose",
    "madeira": "madeira",
    "pesca-aquic": "pesca",
    "min-ferro": "ferro",
    "ouro": "ouro",
    "bauxita": "bauxita",
    "manganes": "manganes",
    "niobio": "niobio",
    "cobre": "cobre",
    "fosfato": "fosfato",
    "carvao-min": "carvao",
    "sal-marinho": "sal-marinho",
    "petroleo": "petroleo",
    "gas-natural": "gas-natural",
    "etanol": "etanol",
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return normalized


def parse_numeric(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if numeric > 0 else None
    raw = str(value).strip()
    if not raw or raw in {"-", "–", "—"}:
        return None
    raw = raw.replace(".", "").replace(",", ".")
    try:
        numeric = float(raw)
    except ValueError:
        return None
    return numeric if numeric > 0 else None


def resolve_city_id(state_code: str, major_city: str, city_ids: set[str]) -> str:
    if (state_code, major_city) in SPECIAL_CITY_IDS:
        return SPECIAL_CITY_IDS[(state_code, major_city)]
    candidate = f"{slugify(state_code)}-{slugify(major_city)}"
    if candidate in city_ids:
        return candidate
    raise KeyError(f"Cidade sem mapeamento no catalogo: {state_code} / {major_city} -> {candidate}")


def header_to_product_id(header_value: str, product_ids: set[str]) -> str:
    normalized = slugify(str(header_value).split("\n", maxsplit=1)[0])
    product_id = HEADER_PRODUCT_IDS.get(normalized)
    if not product_id:
        raise KeyError(f"Cabecalho de produto sem mapeamento: {header_value!r} -> {normalized}")
    if product_id not in product_ids:
        raise KeyError(f"Produto mapeado nao existe no catalogo: {product_id}")
    return product_id


def main() -> None:
    city_catalog = json.loads(CITY_CATALOG_PATH.read_text(encoding="utf-8"))
    product_catalog = json.loads(PRODUCT_CATALOG_PATH.read_text(encoding="utf-8"))
    city_ids = {item["id"] for item in city_catalog}
    product_ids = {item["id"] for item in product_catalog}

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    product_columns: dict[int, str] = {}
    for column_index in range(FIRST_PRODUCT_COLUMN, LAST_PRODUCT_COLUMN + 1):
        header_value = worksheet.cell(row=HEADER_ROW, column=column_index).value
        product_columns[column_index] = header_to_product_id(header_value, product_ids)

    items: list[dict[str, object]] = []
    seen_keys: set[tuple[str, str]] = set()

    for row_index in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        state_code = worksheet.cell(row=row_index, column=1).value
        major_city = worksheet.cell(row=row_index, column=4).value
        if not state_code or not major_city:
            continue

        city_id = resolve_city_id(str(state_code).strip().upper(), str(major_city).strip(), city_ids)
        for column_index, product_id in product_columns.items():
            numeric_value = parse_numeric(worksheet.cell(row=row_index, column=column_index).value)
            if numeric_value is None:
                continue
            matrix_key = (city_id, product_id)
            if matrix_key in seen_keys:
                raise ValueError(f"Item duplicado na matriz de demanda: {matrix_key}")
            seen_keys.add(matrix_key)
            items.append({
                "id": f"city_product_{city_id}_{product_id}",
                "city_id": city_id,
                "product_id": product_id,
                "value": numeric_value,
            })

    OUTPUT_PATH.write_text(json.dumps(items, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"{len(items)} itens salvos em {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
